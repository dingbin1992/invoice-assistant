#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报销封面生成器
基于xlsx模板生成报销封面，然后转为PDF
"""

import os
import sys
import json
import subprocess
from openpyxl import load_workbook

def amount_to_chinese_words(amount):
    """金额转换为中文大写"""
    if amount == 0.0:
        return "零元整"
    
    amount = round(amount * 100)
    jiao = (amount // 10) % 10  # 角
    fen = amount % 10  # 分
    yuan = amount // 100  # 元
    
    units = ["", "拾", "佰", "仟", "万", "拾", "佰", "仟", "亿"]
    digits = ["零", "壹", "贰", "叁", "肆", "伍", "陆", "柒", "捌", "玖"]
    
    result = ""
    yuan_str = str(yuan)
    length = len(yuan_str)
    
    for i, c in enumerate(yuan_str):
        digit = int(c)
        unit_idx = length - i - 1
        if digit != 0:
            result += digits[digit]
            result += units[unit_idx]
        elif result and not result.endswith("零"):
            result += "零"
    
    result = result.rstrip("零")
    if result:
        result += "元"
    
    if jiao != 0 or fen != 0:
        if jiao != 0:
            result += digits[jiao]
            result += "角"
        if fen != 0:
            result += digits[fen]
            result += "分"
        else:
            result += "整"
    else:
        result += "整"
    
    return result

def generate_cover(invoices, output_dir, owner, buyer, output_format='both'):
    """基于xlsx模板生成报销封面，根据output_format生成xlsx/pdf/两种格式"""
    # 获取脚本所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # 上一级目录（和template同级）
    parent_dir = os.path.dirname(script_dir)
    
    # 查找模板文件（在上一级目录的template子目录中）
    template_filename = "费用报销审批单模板.xlsx"
    template_path = os.path.join(parent_dir, "template", template_filename)
    if not os.path.exists(template_path):
        # 回退：在脚本同级目录查找
        template_path = os.path.join(script_dir, template_filename)
    if not os.path.exists(template_path):
        print(f"错误: 找不到模板文件 {template_filename}")
        return []
    
    wb = load_workbook(template_path)
    ws = wb['Sheet1']
    
    # 报销部门留空（打印后手写签字）
    ws['A2'] = "报销部门："
    
    # 按报销类别分组计算金额
    category_amounts = {}
    total_amount = 0.0
    
    for inv in invoices:
        if inv.get('is_invoice_pdf') and inv.get('amount'):
            try:
                amount = float(inv['amount'])
                category = inv.get('category', '未分类')
                if category not in category_amounts:
                    category_amounts[category] = 0.0
                category_amounts[category] += amount
                total_amount += amount
            except ValueError:
                pass
    
    # 填充表格数据（最多5行）
    row_map = {1: 'A5', 2: 'A6', 3: 'A7', 4: 'A8', 5: 'A9'}
    amount_map = {1: 'D5', 2: 'D6', 3: 'D7', 4: 'D8', 5: 'D9'}
    
    row_count = 0
    for category, amount in category_amounts.items():
        if row_count >= 5:
            break
        ws[row_map[row_count + 1]] = category
        cell = ws[amount_map[row_count + 1]]
        cell.value = round(amount, 2)
        cell.number_format = '0.00'
        row_count += 1
    
    # 填充合计
    ws['A10'] = '合计'
    cell = ws['D10']
    cell.value = round(total_amount, 2)
    cell.number_format = '0.00'
    
    # 填充总计人民币(大写)
    ws['A11'] = '总计人民币(大写)'
    ws['D11'] = amount_to_chinese_words(total_amount)
    
    # 设置打印区域为A-I列，页边距窄一点
    from openpyxl.worksheet.page import PageMargins
    ws.print_area = 'A1:I12'
    ws.page_margins = PageMargins(
        left=0.4,    # 左边距
        right=0.4,   # 右边距
        top=0.5,     # 上边距
        bottom=0.5,  # 下边距
        header=0.3,  # 页眉
        footer=0.3   # 页脚
    )
    ws.page_setup.orientation = 'portrait'  # 纵向
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # A4纸
    ws.sheet_properties.pageSetUpPr.fitToPage = True  # 适应页面
    ws.page_setup.fitToWidth = 1  # 宽度适应1页
    ws.page_setup.fitToHeight = 1  # 高度适应1页
    
    # 生成文件名
    base_name = f"费用报销审批单_{owner}_{buyer}"

    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)

    output_files = []
    xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")

    # 总是先保存xlsx文件（PDF转换需要它）
    wb.save(xlsx_path)

    # 根据output_format决定生成哪些文件
    if output_format in ['xlsx', 'both']:
        print(f"已生成xlsx文件: {xlsx_path}")
        output_files.append(xlsx_path)

    if output_format in ['pdf', 'both']:
        # 尝试用Excel将xlsx转为PDF
        pdf_path = os.path.join(output_dir, f"{base_name}.pdf")
        try:
            convert_xlsx_to_pdf(xlsx_path, pdf_path)
            if os.path.exists(pdf_path):
                print(f"已生成PDF文件: {pdf_path}")
                output_files.append(pdf_path)
        except Exception as e:
            print(f"PDF转换失败: {e}")

    # 如果只需要PDF，删除xlsx文件
    if output_format == 'pdf' and os.path.exists(xlsx_path):
        os.remove(xlsx_path)

    return output_files

def convert_xlsx_to_pdf(xlsx_path, pdf_path):
    """使用Excel将xlsx转换为PDF"""
    import win32com.client
    
    excel = None
    wb_excel = None
    
    try:
        # 启动Excel
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # 打开xlsx文件
        abs_xlsx_path = os.path.abspath(xlsx_path)
        wb_excel = excel.Workbooks.Open(abs_xlsx_path)
        
        # 导出为PDF
        abs_pdf_path = os.path.abspath(pdf_path)
        wb_excel.ExportAsFixedFormat(
            Type=0,  # 0 = xlTypePDF
            Filename=abs_pdf_path,
            Quality=0,  # 0 = xlQualityStandard
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False
        )
        
        print(f"PDF转换成功: {pdf_path}")
        
    finally:
        # 关闭工作簿
        if wb_excel:
            wb_excel.Close(SaveChanges=False)
        # 退出Excel
        if excel:
            excel.Quit()
            excel = None

def main():
    if len(sys.argv) < 5:
        print("用法: python generate_cover.py <invoices_json> <output_dir> <owner> <buyer> [output_format]")
        print("output_format: xlsx, pdf, both (默认both)")
        return

    invoices_json_path = sys.argv[1]
    output_dir = sys.argv[2]
    owner = sys.argv[3]
    buyer = sys.argv[4]
    output_format = sys.argv[5] if len(sys.argv) > 5 else 'both'

    # 读取发票数据
    with open(invoices_json_path, 'r', encoding='utf-8') as f:
        invoices = json.load(f)

    # 生成报销封面
    output_files = generate_cover(invoices, output_dir, owner, buyer, output_format)

    print(f"\n生成完成:")
    for f in output_files:
        print(f"  {f}")

if __name__ == "__main__":
    main()
