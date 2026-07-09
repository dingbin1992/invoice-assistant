#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
费用台账生成器
基于xlsx模板生成费用台账，然后转为PDF
"""

import os
import sys
import json
import subprocess
from openpyxl import load_workbook

# 报销类别到列的映射
CATEGORY_COLUMN_MAP = {
    '加油费': {'amount': 'C', 'ticket': 'D'},
    '主干道费': {'amount': 'E', 'ticket': 'F'},
    '住宿费': {'amount': 'G', 'ticket': 'H'},
    '餐饮费': {'amount': 'I', 'ticket': 'J'},
    '其他费用': {'amount': 'K', 'ticket': 'L'},
}

def generate_ledger(invoices, output_dir):
    """基于xlsx模板生成费用台账，然后转为PDF"""
    # 获取脚本所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # 上一级目录（和template同级）
    parent_dir = os.path.dirname(script_dir)
    
    # 查找模板文件
    template_filename = "费用台账模板.xlsx"
    template_path = os.path.join(parent_dir, "template", template_filename)
    if not os.path.exists(template_path):
        # 回退：在脚本同级目录查找
        template_path = os.path.join(script_dir, template_filename)
    if not os.path.exists(template_path):
        print(f"错误: 找不到模板文件 {template_filename}")
        return []
    
    wb = load_workbook(template_path)
    ws = wb['Sheet1']
    
    # 按报销人+购买方分组
    groups = {}
    for inv in invoices:
        if not inv.get('is_invoice_pdf'):
            continue
        owner = inv.get('owner', '未分组')
        buyer = inv.get('buyer', '未知购买方')
        key = f"{owner}_{buyer}"
        if key not in groups:
            groups[key] = {
                'owner': owner,
                'buyer': buyer,
                'invoices': []
            }
        groups[key]['invoices'].append(inv)
    
    # 生成文件名
    base_name = "费用台账"
    
    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)
    
    output_files = []
    
    # 为每个报销人+购买方生成一个台账
    row_start = 3  # 数据从第3行开始
    for idx, (key, group) in enumerate(groups.items()):
        row = row_start + idx
        
        # 填充姓名和区域公司
        ws[f'A{row}'] = group['owner']
        ws[f'B{row}'] = group['buyer']
        
        # 按报销类别统计金额和票号
        category_data = {}
        for inv in group['invoices']:
            category = inv.get('category', '其他费用')
            amount = 0
            try:
                amount = float(inv.get('amount', 0))
            except:
                pass
            invoice_no = inv.get('invoice_no', '')
            
            if category not in category_data:
                category_data[category] = {'total': 0, 'tickets': []}
            category_data[category]['total'] += amount
            if invoice_no:
                category_data[category]['tickets'].append(invoice_no)
        
        # 填充各类别的金额和票号
        total_amount = 0
        for category, col_map in CATEGORY_COLUMN_MAP.items():
            if category in category_data:
                data = category_data[category]
                ws[f"{col_map['amount']}{row}"] = round(data['total'], 2)
                ws[f"{col_map['amount']}{row}"].number_format = '0.00'
                ws[f"{col_map['ticket']}{row}"] = '、'.join(data['tickets'])
                total_amount += data['total']
            else:
                ws[f"{col_map['amount']}{row}"] = 0
                ws[f"{col_map['amount']}{row}"].number_format = '0.00'
                ws[f"{col_map['ticket']}{row}"] = ''
        
        # 填充合计
        ws[f'M{row}'] = round(total_amount, 2)
        ws[f'M{row}'].number_format = '0.00'

    # 计算最后一行
    last_row = row_start + len(groups) - 1
    if last_row < row_start:
        last_row = row_start

    # 设置行高自适应（根据内容自动调整）
    for row in ws.iter_rows(min_row=1, max_row=last_row):
        max_line_count = 1
        for cell in row:
            if cell.value:
                # 估算行数（根据内容长度）
                cell_value = str(cell.value)
                # 计算需要的行数（假设每行约30个字符）
                lines = max(1, len(cell_value) // 30 + 1)
                max_line_count = max(max_line_count, lines)
        # 设置行高（每行约15点）
        ws.row_dimensions[row[0].row].height = max_line_count * 15

    # 设置打印区域为A-M列
    from openpyxl.worksheet.page import PageMargins
    ws.print_area = f'A1:M{last_row}'
    ws.page_margins = PageMargins(
        left=0.4,
        right=0.4,
        top=0.5,
        bottom=0.5,
        header=0.3,
        footer=0.3
    )
    ws.page_setup.orientation = 'landscape'  # 横向
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # A4纸
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # 保存xlsx文件
    xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")
    wb.save(xlsx_path)
    print(f"已生成xlsx文件: {xlsx_path}")
    output_files.append(xlsx_path)

    return output_files

def main():
    if len(sys.argv) < 2:
        print("用法: python generate_ledger.py <invoices_json> <output_dir>")
        return
    
    invoices_json_path = sys.argv[1]
    output_dir = sys.argv[2]
    
    with open(invoices_json_path, 'r', encoding='utf-8') as f:
        invoices = json.load(f)
    
    output_files = generate_ledger(invoices, output_dir)
    
    print(f"\n生成完成:")
    for f in output_files:
        print(f"  {f}")

if __name__ == "__main__":
    main()
